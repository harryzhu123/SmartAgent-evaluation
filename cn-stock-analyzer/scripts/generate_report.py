"""
A股新闻驱动选股分析 - Excel报告生成脚本

用法：
    python generate_report.py --data analysis_data.json --output report.xlsx

输入 JSON 格式见底部文档字符串。
"""

import json
import sys
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule, CellIsRule
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ── A股配色：红涨绿跌 ──
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")      # 利好/上涨
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")    # 利空/下跌（A股绿=跌）
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")   # 中性/警告
ORANGE_FILL = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")   # 高风险
DARK_RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") # 极高风险

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def apply_header(ws, row, headers, col_widths=None):
    """Apply formatted headers to a row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    if col_widths:
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_body_cell(ws, row, col, value, alignment="left"):
    """Apply formatted body cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.alignment = Alignment(
        horizontal=alignment, vertical="center", wrap_text=True
    )
    cell.border = THIN_BORDER
    return cell


def sentiment_fill(score):
    """Return fill color based on sentiment score (A股 convention: red=up/good)."""
    if score is None:
        return None
    if score >= 3:
        return RED_FILL       # 强利好 → 红色
    elif score >= 1:
        return PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    elif score <= -3:
        return GREEN_FILL     # 强利空 → 绿色
    elif score <= -1:
        return PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    return None


def risk_fill(level):
    """Return fill color based on risk level."""
    mapping = {
        "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "中": YELLOW_FILL,
        "高": ORANGE_FILL,
        "极高": DARK_RED_FILL,
    }
    return mapping.get(level)


def create_news_sheet(wb, news_list):
    """Sheet 1: 新闻分析汇总"""
    ws = wb.active
    ws.title = "新闻分析"

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value="📰 新闻分析汇总")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    # Timestamp
    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1, value=f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=2, column=1).font = Font(name="微软雅黑", size=9, color="888888")

    # Headers
    headers = ["序号", "新闻标题", "来源", "发布时间", "情绪评分", "影响行业", "核心判断"]
    col_widths = [6, 40, 12, 14, 10, 20, 35]
    apply_header(ws, 4, headers, col_widths)

    # Data
    for i, news in enumerate(news_list, 1):
        row = i + 4
        ws.row_dimensions[row].height = 40

        apply_body_cell(ws, row, 1, i, "center")
        apply_body_cell(ws, row, 2, news.get("title", ""))
        apply_body_cell(ws, row, 3, news.get("source", ""), "center")
        apply_body_cell(ws, row, 4, news.get("time", ""), "center")

        score = news.get("sentiment_score", 0)
        score_text = f"{'+' if score > 0 else ''}{score}"
        score_cell = apply_body_cell(ws, row, 5, score_text, "center")
        fill = sentiment_fill(score)
        if fill:
            score_cell.fill = fill

        apply_body_cell(ws, row, 6, news.get("industries", ""))
        apply_body_cell(ws, row, 7, news.get("summary", ""))

    # Freeze panes
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{len(news_list) + 4}"


def create_picks_sheet(wb, picks):
    """Sheet 2: 推荐标的"""
    ws = wb.create_sheet("推荐标的")

    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value="🎯 推荐标的列表")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = [
        "股票/ETF", "代码", "板块", "推荐逻辑",
        "情绪评分", "时间维度", "风险等级", "建议仓位", "关键风险"
    ]
    col_widths = [16, 10, 14, 35, 10, 10, 10, 12, 30]
    apply_header(ws, 3, headers, col_widths)

    for i, pick in enumerate(picks, 1):
        row = i + 3
        ws.row_dimensions[row].height = 45

        apply_body_cell(ws, row, 1, pick.get("name", ""))
        apply_body_cell(ws, row, 2, pick.get("code", ""), "center")
        apply_body_cell(ws, row, 3, pick.get("sector", ""), "center")
        apply_body_cell(ws, row, 4, pick.get("reason", ""))

        score = pick.get("sentiment_score", 0)
        score_text = f"{'+' if score > 0 else ''}{score}"
        score_cell = apply_body_cell(ws, row, 5, score_text, "center")
        fill = sentiment_fill(score)
        if fill:
            score_cell.fill = fill

        apply_body_cell(ws, row, 6, pick.get("timeframe", ""), "center")

        risk = pick.get("risk_level", "")
        risk_cell = apply_body_cell(ws, row, 7, risk, "center")
        rfill = risk_fill(risk)
        if rfill:
            risk_cell.fill = rfill

        apply_body_cell(ws, row, 8, pick.get("position", ""), "center")
        apply_body_cell(ws, row, 9, pick.get("risk_note", ""))

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{len(picks) + 3}"


def create_sector_sheet(wb, sectors):
    """Sheet 3: 板块资金流向"""
    if not sectors:
        return

    ws = wb.create_sheet("板块资金流向")

    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value="💰 板块资金流向")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = ["板块名称", "今日涨跌%", "资金净流入(亿)", "5日资金流向", "相关新闻"]
    col_widths = [16, 12, 14, 14, 35]
    apply_header(ws, 3, headers, col_widths)

    for i, sector in enumerate(sectors, 1):
        row = i + 3
        ws.row_dimensions[row].height = 35

        apply_body_cell(ws, row, 1, sector.get("name", ""))

        change = sector.get("change_pct", 0)
        change_cell = apply_body_cell(ws, row, 2, f"{change:+.2f}%", "center")
        if change > 0:
            change_cell.font = Font(name="微软雅黑", size=10, color="FF0000")  # 红=涨
        elif change < 0:
            change_cell.font = Font(name="微软雅黑", size=10, color="00AA00")  # 绿=跌

        flow = sector.get("net_flow", 0)
        flow_cell = apply_body_cell(ws, row, 3, f"{flow:+.2f}", "center")
        if flow > 0:
            flow_cell.font = Font(name="微软雅黑", size=10, color="FF0000")
        elif flow < 0:
            flow_cell.font = Font(name="微软雅黑", size=10, color="00AA00")

        apply_body_cell(ws, row, 4, sector.get("flow_5d", ""), "center")
        apply_body_cell(ws, row, 5, sector.get("related_news", ""))

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{len(sectors) + 3}"


def create_disclaimer_sheet(wb):
    """Sheet 4: 免责声明"""
    ws = wb.create_sheet("免责声明")

    ws.merge_cells("A1:A1")
    ws.cell(row=1, column=1, value="⚠️ 免责声明").font = TITLE_FONT
    ws.column_dimensions["A"].width = 80

    disclaimers = [
        "",
        "1. 本报告由 AI（Claude）基于公开新闻信息和逻辑推理自动生成，仅供参考。",
        "2. 本报告不构成任何投资建议、投资承诺或投资担保。",
        "3. 股市有风险，投资需谨慎。过往表现不代表未来收益。",
        "4. 报告中的行情数据可能存在延迟或不准确，请以官方交易所数据为准。",
        "5. 情绪评分和风险评估基于 AI 分析，存在局限性和不确定性。",
        "6. 仓位建议仅为参考，请结合个人风险承受能力和投资经验做出独立判断。",
        "7. 产业链映射和标的推荐基于公开信息，可能不完整或存在偏差。",
        "",
        "使用本报告即表示您理解并接受以上声明。",
    ]

    for i, line in enumerate(disclaimers, 2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name="微软雅黑", size=11)
        cell.alignment = Alignment(wrap_text=True)


def generate_report(data, output_path):
    """Main entry: generate the full Excel report."""
    wb = Workbook()

    create_news_sheet(wb, data.get("news", []))
    create_picks_sheet(wb, data.get("picks", []))
    create_sector_sheet(wb, data.get("sectors", []))
    create_disclaimer_sheet(wb)

    wb.save(output_path)
    print(f"报告已生成：{output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成A股新闻选股分析报告")
    parser.add_argument("--data", required=True, help="分析数据 JSON 文件路径")
    parser.add_argument("--output", default="stock_analysis_report.xlsx", help="输出文件路径")
    args = parser.parse_args()

    with open(args.data, "r", encoding="utf-8") as f:
        data = json.load(f)

    generate_report(data, args.output)


"""
输入 JSON 格式示例：

{
  "news": [
    {
      "title": "国务院发布新能源汽车下乡支持政策",
      "source": "新华社",
      "time": "2025-03-05",
      "sentiment_score": 4,
      "industries": "新能源汽车、锂电池",
      "summary": "明确利好，财政补贴+购置税减免延续"
    }
  ],
  "picks": [
    {
      "name": "宁德时代",
      "code": "300750",
      "sector": "锂电池",
      "reason": "新能源车下乡直接拉动动力电池需求，公司市占率第一",
      "sentiment_score": 4,
      "timeframe": "中线",
      "risk_level": "中",
      "position": "5%-8%",
      "risk_note": "锂价波动影响毛利率，需关注原材料走势"
    }
  ],
  "sectors": [
    {
      "name": "新能源车板块",
      "change_pct": 2.35,
      "net_flow": 15.8,
      "flow_5d": "持续流入",
      "related_news": "新能源汽车下乡政策"
    }
  ]
}
"""
